"""
Sync the commercial-master product catalog from Google Sheets
→ campaigns.product_catalog

Sources (two public Google Sheets, read via the XLSX export endpoint):
  MASTER  "MEAMA Commercial Master 2026"  — per-SKU COGS, price, margin, discount logic
            tabs: Espresso Format, Multicapsule Format, Classic Coffee,
                  Machines & Accessories, Inputs
  BIBLE   "MEAMA Core Product Bible"      — identity, sensory, preparation
            tabs: 01 · Capsule Products, 02 · Classic Coffee, 04 · Preparation Guide

The per-SKU economics tabs are the spine; the Bible enriches names / sensory /
preparation. SKUs present in the economics tabs but absent from the Bible are
marked 'discontinued' (they still carry real COGS history).

Margin / max-safe-discount are NOT computed here — the backend derives them from
(price_per_unit, total_cogs) via business_rules.py.

Usage:
    python3 pipeline/sync_product_catalog.py            # read public sheets + upsert
    python3 pipeline/sync_product_catalog.py --dry-run  # parse + print, no DB write

Requires in .env:
    SUPABASE_URL
    SUPABASE_SERVICE_ROLE_KEY
Optional overrides:
    MEAMA_MASTER_SHEET_ID   (default below)
    MEAMA_BIBLE_SHEET_ID    (default below)
"""

from __future__ import annotations

import argparse
import io
import os
import re
from collections import defaultdict
from datetime import datetime, timezone

import openpyxl
import requests
from dotenv import load_dotenv

load_dotenv()

MASTER_SHEET_ID = os.environ.get(
    "MEAMA_MASTER_SHEET_ID", "14_Ao5v13u8rm8BpA8Bb3mZ7NOeQQG8xuByfWhgoIZxE"
)
BIBLE_SHEET_ID = os.environ.get(
    "MEAMA_BIBLE_SHEET_ID", "1CpgYks5xXyaKMMW6R2Tfp1ozBDMOHVbMaGZTu3ctnOY"
)

# A value looks like a SKU when it carries a known prefix or a code-with-digits shape.
SKU_RE = re.compile(r"^(cap\d+|tea\d+|Mix|tb-|bag-|can-|ch-|ab-|mcm|gc|dwg|K51|Pinta)", re.IGNORECASE)


# ── Sheet fetch ──────────────────────────────────────────────────────────────

def load_workbook(sheet_id: str) -> openpyxl.Workbook:
    """Download a public Google Sheet as XLSX (raw, unrounded cell values)."""
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    resp = requests.get(url, timeout=60)
    resp.raise_for_status()
    return openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True)


def norm(s) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip()).lower()


def num(v):
    """Coerce a cell to float, tolerating '29,840.00 ₾' / '₾1.50' / '70%' style strings."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace("₾", "").replace(",", "").replace("%", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def looks_like_sku(v) -> bool:
    return bool(v) and bool(SKU_RE.match(str(v).strip()))


# ── Economics parsing (Master) ───────────────────────────────────────────────

def parse_economics_tab(ws, product_type: str, price_col: str) -> list[dict]:
    """Walk a tab with repeating header bands + section labels, emitting one row
    per SKU. `price_col` is the header that holds the per-unit price
    ('price per cap' for capsules, 'selling price' for classic/machines)."""
    # SKU column may hold plain machine/accessory names (e.g. 'Versatile'),
    # so a data row is identified by a non-empty SKU cell that is not a band
    # label / 'Averages' summary, plus a numeric per-unit price.
    stop = {"", "sku", "averages", "category", "subcateogory", "subcategory"}
    rows: list[dict] = []
    header: dict[str, int] = {}
    for r in ws.iter_rows(values_only=True):
        cells = list(r)
        labels = [norm(c) for c in cells]
        # (Re)capture a header band whenever we see the SKU/Product Name columns.
        if "sku" in labels and ("product name" in labels or "product" in labels):
            header = {norm(c): i for i, c in enumerate(cells) if norm(c)}
            continue
        if not header:
            continue
        sku_i = header.get("sku")
        if sku_i is None or sku_i >= len(cells):
            continue
        sku = cells[sku_i]
        if sku is None or norm(sku) in stop:
            continue  # skips section labels, 'Averages', blank rows

        def col(*names):
            for n in names:
                i = header.get(n)
                if i is not None and i < len(cells):
                    return cells[i]
            return None

        price_per_unit = num(col(price_col, "price per cap", "selling price"))
        if price_per_unit is None:
            continue  # band labels / spacer rows carry no price
        price_per_pack = num(col("price per pack")) or price_per_unit
        caps = None
        if price_per_unit and price_per_pack and price_per_unit > 0:
            caps = round(price_per_pack / price_per_unit)
        # Accessories leave 'Total COGs' blank — fall back to the base COGs column.
        total_cogs = num(col("total cogs")) or num(col("production cogs", "cogs"))
        rows.append({
            "sku": str(sku).strip(),
            "product_type": product_type,
            "category": (col("category") or "").strip() or None,
            "subcategory": (col("subcateogory", "subcategory") or "").strip() or None,
            "name_en": (col("product name", "product") or "").strip() or None,
            "production_cogs": num(col("production cogs", "cogs")),
            "total_cogs": total_cogs,
            "price_per_pack": price_per_pack,
            "price_per_unit": price_per_unit,
            "caps_per_pack": caps if product_type == "capsule" else None,
            "full_margin": num(col("margin per capsule", "margin")),
            "sales_units": num(col("sales")),
            "current_stock": num(col("current stock")),
            "source_tab": ws.title,
        })
    return rows


def split_machines_accessories(rows: list[dict]) -> None:
    """The 'Machines & Accessories' tab mixes both; tag accessories by category."""
    for row in rows:
        cat = norm(row.get("category"))
        if "accs" in cat or "accessor" in cat:
            row["product_type"] = "accessory"
        else:
            row["product_type"] = "machine"


# ── Bible enrichment ─────────────────────────────────────────────────────────

def parse_bible_capsules(ws) -> dict[str, dict]:
    out: dict[str, dict] = {}
    header: dict[str, int] = {}
    for r in ws.iter_rows(values_only=True):
        cells = list(r)
        labels = [norm(c) for c in cells]
        if "sku" in labels and any("name (english" in l for l in labels):
            header = {norm(c): i for i, c in enumerate(cells) if norm(c)}
            continue
        if not header:
            continue
        sku_i = header.get("sku")
        if sku_i is None or sku_i >= len(cells) or not looks_like_sku(cells[sku_i]):
            continue

        def col(*names):
            for n in names:
                for hk, i in header.items():
                    if hk.startswith(n) and i < len(cells):
                        return cells[i]
            return None

        sku = str(cells[sku_i]).strip()
        out[sku] = {
            "production_name": (col("production name") or "").strip() or None,
            "name_en": (col("name (english") or "").strip() or None,
            "name_ka": (col("name (georgian") or "").strip() or None,
            "format": (col("format") or "").strip() or None,
            "category": (col("category") or "").strip() or None,
            "subcategory": (col("subcategory") or "").strip() or None,
            "status": (col("status") or "").strip().lower() or "live",
            "intensity": num(col("intensity")),
            "bitterness": num(col("bitterness")),
            "caffeine_mg": num(col("caffeine (mg")),
            "flavour_notes": (col("flavour notes") or "").strip() or None,
        }
    return out


def parse_bible_classic(ws) -> dict[str, dict]:
    out: dict[str, dict] = {}
    header: dict[str, int] = {}
    for r in ws.iter_rows(values_only=True):
        cells = list(r)
        labels = [norm(c) for c in cells]
        if "sku" in labels and any("name (english" in l for l in labels):
            header = {norm(c): i for i, c in enumerate(cells) if norm(c)}
            continue
        if not header:
            continue
        sku_i = header.get("sku")
        if sku_i is None or sku_i >= len(cells) or not looks_like_sku(cells[sku_i]):
            continue

        def col(*names):
            for n in names:
                for hk, i in header.items():
                    if hk.startswith(n) and i < len(cells):
                        return cells[i]
            return None

        sku = str(cells[sku_i]).strip()
        out[sku] = {
            "production_name": (col("production name") or "").strip() or None,
            "name_en": (col("name (english") or "").strip() or None,
            "name_ka": (col("name (georgian") or "").strip() or None,
            "status": (col("status") or "").strip().lower() or "live",
            "flavour_notes": (col("flavour notes") or "").strip() or None,
        }
    return out


def parse_bible_prep(ws) -> dict[str, dict]:
    """Aggregate compatible machines per SKU from the Preparation Guide."""
    machines: dict[str, set] = defaultdict(set)
    program: dict[str, str] = {}
    serving: dict[str, str] = {}
    header: dict[str, int] = {}
    for r in ws.iter_rows(values_only=True):
        cells = list(r)
        labels = [norm(c) for c in cells]
        if "sku" in labels and any("compatible machine" in l for l in labels):
            header = {norm(c): i for i, c in enumerate(cells) if norm(c)}
            continue
        if not header:
            continue
        sku_i = header.get("sku")
        if sku_i is None or sku_i >= len(cells) or not looks_like_sku(cells[sku_i]):
            continue

        def col(*names):
            for n in names:
                for hk, i in header.items():
                    if hk.startswith(n) and i < len(cells):
                        return cells[i]
            return None

        sku = str(cells[sku_i]).strip()
        m = (col("compatible machine") or "").strip()
        if m:
            machines[sku].add(m)
        if col("recommended program"):
            program[sku] = str(col("recommended program")).strip()
        if col("serving"):
            serving[sku] = str(col("serving")).strip()
    return {
        sku: {
            "compatible_machines": sorted(machines[sku]) or None,
            "recommended_program": program.get(sku),
            "serving": serving.get(sku),
        }
        for sku in machines
    }


# ── Assembly ─────────────────────────────────────────────────────────────────

def build_rows(master: openpyxl.Workbook, bible: openpyxl.Workbook) -> list[dict]:
    econ: list[dict] = []
    econ += parse_economics_tab(master["Espresso Format"], "capsule", "price per cap")
    econ += parse_economics_tab(master["Multicapsule Format"], "capsule", "price per cap")
    econ += parse_economics_tab(master["Classic Coffee"], "classic_coffee", "selling price")
    ma = parse_economics_tab(master["Machines & Accessories"], "machine", "selling price")
    split_machines_accessories(ma)
    econ += ma

    caps_meta = parse_bible_capsules(bible["01 · Capsule Products"])
    classic_meta = parse_bible_classic(bible["02 · Classic Coffee"])
    prep_meta = parse_bible_prep(bible["04 · Preparation Guide"])
    bible_skus = set(caps_meta) | set(classic_meta)

    # Dedupe economics by SKU (later tab wins; all share the same SKU key).
    by_sku: dict[str, dict] = {}
    for row in econ:
        by_sku[row["sku"]] = {**by_sku.get(row["sku"], {}), **row}

    now = datetime.now(timezone.utc).isoformat()
    rows: list[dict] = []
    for sku, base in by_sku.items():
        meta = caps_meta.get(sku) or classic_meta.get(sku) or {}
        prep = prep_meta.get(sku, {})
        # Bible names/status/sensory override the terse economics labels.
        merged = {**base}
        for k, v in meta.items():
            if v is not None:
                merged[k] = v
        merged.update({k: v for k, v in prep.items() if v is not None})
        # The Bible only governs capsules + classic coffee; a capsule/classic SKU
        # absent from it is a discontinued line (kept for COGS history). Machines
        # and accessories are never in the Bible, so they stay 'live'.
        if (
            merged.get("product_type") in ("capsule", "classic_coffee")
            and sku not in bible_skus
            and not merged.get("status")
        ):
            merged["status"] = "discontinued"
        merged.setdefault("status", "live")
        merged["synced_at"] = now
        merged["raw"] = {
            k: base.get(k) for k in ("full_margin", "sales_units", "current_stock")
        }
        rows.append(merged)
    return rows


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true", help="parse + print, skip DB write")
    args = ap.parse_args()

    print("Fetching Master + Bible workbooks…")
    master = load_workbook(MASTER_SHEET_ID)
    bible = load_workbook(BIBLE_SHEET_ID)

    rows = build_rows(master, bible)
    from collections import Counter
    counts = Counter(r["product_type"] for r in rows)
    print(f"Parsed {len(rows)} SKUs: " + ", ".join(f"{k}={v}" for k, v in sorted(counts.items())))
    missing_cogs = [r["sku"] for r in rows if r.get("total_cogs") is None]
    if missing_cogs:
        print(f"  ⚠ {len(missing_cogs)} SKUs without total_cogs: {missing_cogs[:10]}")

    if args.dry_run:
        for r in rows[:8]:
            print(f"  {r['sku']:<14} {r.get('name_en','?'):<22} "
                  f"₾{r.get('price_per_unit')}/u  COGS ₾{r.get('total_cogs')}  "
                  f"[{r['product_type']}/{r.get('status')}]")
        print("dry-run — nothing written")
        return

    from supabase import create_client
    sb = create_client(os.environ["SUPABASE_URL"], os.environ["SUPABASE_SERVICE_ROLE_KEY"])
    print(f"Upserting {len(rows)} rows to campaigns.product_catalog…")
    for i in range(0, len(rows), 50):
        chunk = rows[i : i + 50]
        sb.schema("campaigns").from_("product_catalog").upsert(
            chunk, on_conflict="sku"
        ).execute()
        print(f"  upserted {i + 1}–{i + len(chunk)}")
    print(f"Done — {len(rows)} SKUs in campaigns.product_catalog")


if __name__ == "__main__":
    main()
